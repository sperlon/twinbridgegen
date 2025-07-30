# Author:  Radek Marik, radek.marik@fel.cvut.cz
# Created: Dec 21, 2023
# Purpose: input data readers

# Copyright (c) 2023, Radek Marik, FEE CTU, Prague
# All rights reserved.
#
# Redistribution and use in source and binary forms, with or without
# modification, are permitted provided that the following conditions are met:
#     * Redistributions of source code must retain the above copyright
#       notice, this list of conditions and the following disclaimer.
#     * Redistributions in binary form must reproduce the above copyright
#       notice, this list of conditions and the following disclaimer in the
#       documentation and/or other materials provided with the distribution.
#     * Neither the name of FEE CTU nor the
#       names of contributors may be used to endorse or promote products
#       derived from this software without specific prior written permission.
#
# THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS "AS IS" AND
# ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE IMPLIED
# WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE
# DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT HOLDER OR CONTRIBUTORS BE LIABLE FOR
# ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL DAMAGES
# (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF SUBSTITUTE GOODS OR SERVICES;
# LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER CAUSED AND
# ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT LIABILITY, OR TORT
# (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY OUT OF THE USE OF THIS
# SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH DAMAGE.

from __future__ import print_function
import numpy as np
import pickle
import pathlib
import openpyxl   # Excel xlsx reader
Set_C = set

# =======================================================================
# ===================== Support Structures ==============================
# =======================================================================

# =======================================================================
# ===================== Support Functions ===============================
# =======================================================================

def PrepareSheetcfg(runCfg, training=True):
  if training:
    inSheetName = runCfg["trainingDataset"]["inSheet"].strip()
    inColOffset = runCfg["trainingDataset"]["inColOffset"]
    outSheetName = runCfg["trainingDataset"]["outSheet"].strip()
    outColOffset = runCfg["trainingDataset"]["outColOffset"]
    sheetCfg = {
      "inSheetName": runCfg["trainingDataset"]["inSheet"].strip(),
      "inColOffset": runCfg["trainingDataset"]["inColOffset"],
      "outSheetName": runCfg["trainingDataset"]["outSheet"].strip(),
      "outColOffset": runCfg["trainingDataset"]["outColOffset"],
      "sheetList": []
    }


# =======================================================================
# ===================== Classes =========================================
# =======================================================================

# =======================================================================
# ===================== Main Functions ==================================
# =======================================================================

def ReadInputData(inPN, outPN, verbose, runCfg, mode="training"):
  #=====================================
  verboseLevel = 3
  wb = openpyxl.load_workbook(inPN)
  if verbose > verboseLevel: print('sheetsN:', inPN, wb.sheetnames)

  if mode == "training":
    inSheetName  = runCfg["trainingDataset"]["inSheet"].strip()
    inColOffset  = runCfg["trainingDataset"]["inColOffset"]
    outSheetName = runCfg["trainingDataset"]["outSheet"].strip()
    outColOffset = runCfg["trainingDataset"]["outColOffset"]
    sheetNames = [inSheetName, outSheetName]
    sheetOffsets = [inColOffset, outColOffset]

  elif mode == "prediction":
    inSheetName = runCfg["predictionDataset"]["inSheet"].strip()
    inColOffset = runCfg["predictionDataset"]["inColOffset"]
    sheetNames = [inSheetName]
    sheetOffsets = [inColOffset]

  elif mode == "inverseAnalysis":
    inSheetName = runCfg["inverseAnalysisDataset"]["inSheet"].strip()
    inColOffset = runCfg["inverseAnalysisDataset"]["inColOffset"]
    sheetNames = [inSheetName]
    sheetOffsets = [inColOffset]

  else:
    raise Exception(f"Invalid reading mode: '{mode}'.")


  if not Set_C(sheetNames) <= Set_C(wb.sheetnames):
    print('FAILURE: required excel sheets not found.', repr(sheetNames), wb.sheetnames)
    exit(1)

  dataPool = {sheetName: {} for sheetName in sheetNames}
  for sheetName, colOffset in zip(sheetNames, sheetOffsets):
    vSheet = wb[sheetName]
    maxC = None
    for i in range(0, vSheet.max_column):
      if vSheet[1][i].value is not None:
        maxC = i

    colSheetNames = [vSheet[1][c].value.replace("\n", " ") for c in range(0, maxC + 1)]
    colNames = ['sample'] + colSheetNames[colOffset:]

    indexRecords = []
    typeRecords  = []
    dataRecords  = []
    for rId in range(3, vSheet.max_row + 1):
      shRow = vSheet[rId]
      if shRow[0].value is None: break
      rowIndex = int(shRow[0].value)
      rowType  = str(shRow[1].value)
      indexRecords.append(rowIndex)
      typeRecords.append(rowType)
      dataRecords.append([float(rowIndex)] + [float(shRow[j].value) for j in range(colOffset, maxC + 1)])
    dataRecords  = np.array(dataRecords)
    indexRecords = np.array(indexRecords, dtype=int)

    np.save(outPN /f'{sheetName}_data.npy', dataRecords)
    np.save(outPN / f'{sheetName}_index.npy', indexRecords)
    with open(outPN /f'{sheetName}_index.pck', 'wb') as fOut:
      pickle.dump(indexRecords, fOut)
      pickle.dump(typeRecords, fOut)
      pickle.dump(colNames, fOut)
    dataPool[sheetName] = {'data': dataRecords, 'index': indexRecords, 'rtype': typeRecords, 'colNames': colNames}

    #Print info about input structure
    if verbose > verboseLevel:
      print('sheet:', sheetName, vSheet.max_row, vSheet.max_column, maxC)
      print('colSheetNames', sheetName, len(colSheetNames), colSheetNames)
      print('colNames', sheetName, len(colNames), colNames)
      print('dataRecords', dataRecords.shape, dataRecords[0])



  return dataPool

# =======================================================================
# ===================== Resource Registration ===========================
# =======================================================================

# =======================================================================
# ===================== Test Functions ==================================
# =======================================================================

  
# =======================================================================
# ===================== MAIN ============================================
# =======================================================================

if __name__ == '__main__':
  pass
  
